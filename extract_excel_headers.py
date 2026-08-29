import sys
import os
try:
    import openpyxl
    files = ['Automation_Test_Report.xlsx', 'Execution_Summary.xlsx', 'Failed_Test_Cases.xlsx', 'Passed_Test_Cases.xlsx', 'Summary_Report.xlsx']
    for f in files:
        filepath = os.path.join('Testcase_reference', f)
        if not os.path.exists(filepath):
            print(f'File not found: {filepath}')
            continue
        wb = openpyxl.load_workbook(filepath, read_only=True)
        print(f'--- {f} ---')
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1):
                headers = [cell.value for cell in row]
                break
            print(f'Sheet: {sheet}, Headers: {headers}')
except Exception as e:
    print('Error:', e)
