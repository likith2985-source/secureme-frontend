import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import datetime
import random

def create_reference_excels(suite_name, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    
    # Styles
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    thin_side = Side(border_style="thin", color="CBD5E0")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def apply_header_style(ws, headers):
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = cell_border

    def append_data(ws, data_rows):
        for row_idx, data in enumerate(data_rows, 2):
            for col_idx, val in enumerate(data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.font = Font(name="Segoe UI", size=10)
                c.border = cell_border

    def auto_adjust_columns(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    # Mock Data
    total_tests = 300
    passed_tests = 300
    failed_tests = 0
    pass_rate = 100.0
    duration = 120
    start_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    end_time = (datetime.datetime.now() + datetime.timedelta(seconds=duration)).strftime("%Y-%m-%d %H:%M:%S")

    test_categories = ["UI", "API", "Security", "Performance", "Integration"]
    mock_test_cases = []
    for i in range(1, total_tests + 1):
        category = random.choice(test_categories)
        test_name = f"Verify {suite_name} functionality case {i}"
        time_taken = round(random.uniform(0.5, 3.5), 2)
        mock_test_cases.append((i, category, test_name, time_taken, "Passed"))

    # 1. Automation_Test_Report.xlsx
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "Summary"
    apply_header_style(ws1, ['Test Suite', 'Total Tests', 'Passed', 'Failed', 'Pass Rate %', 'Duration (sec)', 'Start Time', 'End Time'])
    append_data(ws1, [[f"{suite_name} Suite", total_tests, passed_tests, failed_tests, pass_rate, duration, start_time, end_time]])
    auto_adjust_columns(ws1)

    ws1_p = wb1.create_sheet("Passed Tests")
    apply_header_style(ws1_p, ['No.', 'Category', 'Test Name', 'Time (sec)', 'Status'])
    append_data(ws1_p, mock_test_cases)
    auto_adjust_columns(ws1_p)

    ws1_f = wb1.create_sheet("Failed Tests")
    apply_header_style(ws1_f, ['No.', 'Category', 'Test Name', 'Error'])
    # Empty as no failures
    auto_adjust_columns(ws1_f)

    ws1_l = wb1.create_sheet("Execution Log")
    apply_header_style(ws1_l, ['Timestamp', 'Level', 'Message'])
    append_data(ws1_l, [[start_time, "INFO", f"Started {suite_name} test execution"], [end_time, "INFO", "Finished test execution successfully"]])
    auto_adjust_columns(ws1_l)

    ws1_d = wb1.create_sheet("Test Details")
    apply_header_style(ws1_d, ['No.', 'Category', 'Test Name', 'Status', 'Error Details'])
    append_data(ws1_d, [[t[0], t[1], t[2], t[4], "N/A"] for t in mock_test_cases])
    auto_adjust_columns(ws1_d)

    wb1.save(os.path.join(output_dir, "Automation_Test_Report.xlsx"))

    # 2. Execution_Summary.xlsx
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Summary"
    apply_header_style(ws2, ['Metric', 'Value'])
    append_data(ws2, [
        ["Total Executed", total_tests],
        ["Total Passed", passed_tests],
        ["Total Failed", failed_tests],
        ["Overall Pass Rate", f"{pass_rate}%"],
        ["Total Execution Time", f"{duration} seconds"]
    ])
    auto_adjust_columns(ws2)
    wb2.save(os.path.join(output_dir, "Execution_Summary.xlsx"))

    # 3. Failed_Test_Cases.xlsx
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "Failed Tests"
    apply_header_style(ws3, ['No.', 'Category', 'Test Name', 'Error'])
    auto_adjust_columns(ws3)
    wb3.save(os.path.join(output_dir, "Failed_Test_Cases.xlsx"))

    # 4. Passed_Test_Cases.xlsx
    wb4 = openpyxl.Workbook()
    ws4 = wb4.active
    ws4.title = "Passed Tests"
    apply_header_style(ws4, ['No.', 'Category', 'Test Name', 'Time (sec)', 'Status'])
    append_data(ws4, mock_test_cases)
    auto_adjust_columns(ws4)
    wb4.save(os.path.join(output_dir, "Passed_Test_Cases.xlsx"))

    # 5. Summary_Report.xlsx
    wb5 = openpyxl.Workbook()
    ws5 = wb5.active
    ws5.title = "Summary"
    apply_header_style(ws5, ['Test Suite', 'Total Tests', 'Passed', 'Failed', 'Pass Rate %', 'Duration (sec)', 'Start Time', 'End Time'])
    append_data(ws5, [[f"{suite_name} Combined Suite", total_tests, passed_tests, failed_tests, pass_rate, duration, start_time, end_time]])
    auto_adjust_columns(ws5)
    wb5.save(os.path.join(output_dir, "Summary_Report.xlsx"))

    print(f"[SUCCESS] Generated 5 Excel reports in '{output_dir}' directory for suite '{suite_name}'.")

if __name__ == "__main__":
    suite_name = sys.argv[1] if len(sys.argv) > 1 else "Generic"
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "."
    create_reference_excels(suite_name, output_dir)
